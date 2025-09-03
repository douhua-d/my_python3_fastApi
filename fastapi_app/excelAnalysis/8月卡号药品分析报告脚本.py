import pandas as pd
from openpyxl import load_workbook

# Excel 文件路径（与脚本同层级）
files = {
    'August': '可力洛8月.xlsx',
    'July': '可力洛7月.xlsx',
    'June': '可力洛6月.xlsx',
    'May': '可力洛5月.xlsx',
    'April': '可力洛4月.xlsx',
    'March': '可力洛3月.xlsx',
    'February': '可力洛2月.xlsx',
    'January': '可力洛1月.xlsx'
}

# 读取每个月的数据（包含所有列）
def read_excel_data(file_path):
    try:
        df = pd.read_excel(file_path)
        return df
    except FileNotFoundError:
        print(f"文件 {file_path} 不存在，跳过...")
        return pd.DataFrame()

# 提取所有月份的数据
dataframes = {month: read_excel_data(path) for month, path in files.items() if not read_excel_data(path).empty}

# 从1-7月数据中提取出所有的就诊卡号
previous_cards = set()
for month in ['January', 'February', 'March', 'April', 'May', 'June', 'July']:
    if month in dataframes and not dataframes[month].empty:
        previous_cards.update(dataframes[month]['就诊卡号'].astype(str).tolist())

# 获取8月数据
if 'August' not in dataframes or dataframes['August'].empty:
    print("8月数据文件不存在或为空，请检查文件路径")
    exit()

august_data = dataframes['August']

# 将8月数据中就诊卡号不在之前月份的行筛选出来，即新增卡号
new_cards_data = august_data[~august_data['就诊卡号'].astype(str).isin(previous_cards)]

# 按医生分组统计8月份总卡号数量、新增卡号数量和药品数量
doctor_stats = []
for doctor, group in august_data.groupby('医生'):
    # 获取该医生的总卡号数
    total_cards = len(group)
    
    # 获取该医生的新增卡号数
    new_cards = sum(~group['就诊卡号'].astype(str).isin(previous_cards))
    
    # 获取该医生的科室/单位（取最常见的）
    unit = group['科室'].value_counts().index[0] if '科室' in group.columns else "未知"
    
    # 获取该医生的药品和数量统计
    total_medicine_qty = group['数量'].sum() if '数量' in group.columns else 0
    avg_medicine_per_card = round(total_medicine_qty / total_cards, 2) if total_cards > 0 else 0
    
    # 获取药品名称和单位（假设列名为"药品"和"单位"，根据实际调整）
    main_medicine = group['药品'].value_counts().index[0] if '药品' in group.columns else "未知"
    medicine_unit = group['单位'].value_counts().index[0] if '单位' in group.columns else "未知"
    
    # 添加到统计列表
    doctor_stats.append({
        '医生': doctor,
        '科室': unit,
        '8月总卡号数量': total_cards,
        '8月新增卡号数量': new_cards,
        '新增占比': f"{(new_cards / total_cards * 100):.2f}%" if total_cards > 0 else "0%",
        '主要药品': main_medicine,
        '单位': medicine_unit,
        '总药品数量': total_medicine_qty,
        '平均每卡药品数': avg_medicine_per_card
    })

# 转换为DataFrame并按新增数量排序
doctor_stats_df = pd.DataFrame(doctor_stats).sort_values('8月新增卡号数量', ascending=False)

# 药品统计分析
medicine_stats = []

# 全部药品统计
if '药品' in august_data.columns and '数量' in august_data.columns:
    # 按药品分组统计
    for medicine, group in august_data.groupby('药品'):
        # 总数量
        total_qty = group['数量'].sum()
        # 使用该药品的医生数
        doctors_count = group['医生'].nunique()
        # 使用该药品的患者数
        patients_count = group['就诊卡号'].nunique()
        # 该药品新增患者数
        new_patients = group[~group['就诊卡号'].astype(str).isin(previous_cards)]['就诊卡号'].nunique()
        
        medicine_stats.append({
            '药品': medicine,
            '单位': group['单位'].value_counts().index[0] if '单位' in group.columns else "未知",
            '总数量': total_qty,
            '使用医生数': doctors_count,
            '使用患者数': patients_count,
            '新增患者数': new_patients,
            '新增患者占比': f"{(new_patients / patients_count * 100):.2f}%" if patients_count > 0 else "0%"
        })

medicine_stats_df = pd.DataFrame(medicine_stats).sort_values('总数量', ascending=False)

# 月度对比分析
monthly_comparison = []
for month in ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August']:
    if month in dataframes and not dataframes[month].empty:
        data = dataframes[month]
        total_cards = data['就诊卡号'].nunique()
        total_medicine_qty = data['数量'].sum() if '数量' in data.columns else 0
        avg_medicine_per_card = round(total_medicine_qty / total_cards, 2) if total_cards > 0 else 0
        
        monthly_comparison.append({
            '月份': month,
            '总卡号数': total_cards,
            '总药品数量': total_medicine_qty,
            '平均每卡药品数': avg_medicine_per_card,
            '医生数': data['医生'].nunique(),
            '药品种类数': data['药品'].nunique() if '药品' in data.columns else 0
        })

monthly_comparison_df = pd.DataFrame(monthly_comparison)

# 创建一个ExcelWriter对象，用于输出Excel文件
with pd.ExcelWriter('8月卡号药品分析报告.xlsx', engine='openpyxl') as writer:
    # 1. 医生统计表（现在包含药品和单位信息）
    doctor_stats_df.to_excel(writer, sheet_name='医生统计', index=False)
    
    # 2. 药品统计表
    if not medicine_stats_df.empty:
        medicine_stats_df.to_excel(writer, sheet_name='药品统计', index=False)
    
    # 3. 月度对比分析
    monthly_comparison_df.to_excel(writer, sheet_name='月度对比', index=False)
    
    # 4. 全部8月数据(可筛选)
    august_data.to_excel(writer, sheet_name='8月全部数据', index=False)
    
    # 5. 全部新增卡号数据(可筛选)
    new_cards_data.to_excel(writer, sheet_name='8月新增卡号', index=False)
    
    # 6. 按医生分组的全部数据
    for doctor, group in august_data.groupby('医生'):
        sheet_name = f"{doctor}-全部"[:31]
        group.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # 7. 主要药品分析（取前10种药品）
    if '药品' in august_data.columns:
        top_medicines = august_data['药品'].value_counts().head(10)
        top_med_df = pd.DataFrame({'药品': top_medicines.index, '出现次数': top_medicines.values})
        top_med_df.to_excel(writer, sheet_name='主要药品TOP10', index=False)

# 添加自动筛选功能
wb = load_workbook('8月卡号药品分析报告.xlsx')
for sheet_name in ['医生统计', '药品统计', '月度对比', '8月全部数据', '8月新增卡号', '主要药品TOP10']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.auto_filter.ref = ws.dimensions

# 保存工作簿
wb.save('8月卡号药品分析报告.xlsx')

print('8月卡号药品分析报告已导出为 8月卡号药品分析报告.xlsx')
print(f"8月总卡号数: {august_data['就诊卡号'].nunique()}")
print(f"8月新增卡号数: {len(new_cards_data)}")
print(f"8月新增占比: {(len(new_cards_data) / august_data['就诊卡号'].nunique() * 100):.2f}%")
