import pandas as pd
from openpyxl import load_workbook

# Excel 文件路径（与脚本同层级）
files = {
    'June': '可力洛6月.xlsx',
    'May': '可力洛5月.xlsx',
    'April': '可力洛4月.xlsx',
    'March': '可力洛3月.xlsx',
    'February': '可力洛2月.xlsx',
    'January': '可力洛1月.xlsx'
}

# 读取每个月的数据（包含所有列）
def read_excel_data(file_path):
    try:
        df = pd.read_excel(file_path)
        return df
    except FileNotFoundError:
        print(f"文件 {file_path} 不存在，跳过...")
        return pd.DataFrame()

# 提取所有月份的数据
dataframes = {month: read_excel_data(path) for month, path in files.items() if not read_excel_data(path).empty}

# 从1-5月数据中提取出所有的就诊卡号
previous_cards = set()
for month in ['January', 'February', 'March', 'April', 'May']:
    if month in dataframes and not dataframes[month].empty:
        previous_cards.update(dataframes[month]['就诊卡号'].astype(str).tolist())

# 获取6月数据
if 'June' not in dataframes or dataframes['June'].empty:
    print("6月数据文件不存在或为空，请检查文件路径")
    exit()

june_data = dataframes['June']

# 将6月数据中就诊卡号不在之前月份的行筛选出来，即新增卡号
new_cards_data = june_data[~june_data['就诊卡号'].astype(str).isin(previous_cards)]

# 筛选单位为"盒"的新增卡号数据
new_cards_box_data = new_cards_data[new_cards_data['单位'] == '盒']

# 按医生分组统计新增盒装药品数据
doctor_box_stats = []
for doctor, group in new_cards_box_data.groupby('医生'):
    # 获取该医生的新增卡号数
    new_cards_count = group['就诊卡号'].nunique()
    
    # 获取该医生的科室（取最常见的）
    unit = group['科室'].value_counts().index[0] if '科室' in group.columns else "未知"
    
    # 获取该医生的药品和数量统计
    total_medicine_qty = group['数量'].sum() if '数量' in group.columns else 0
    
    # 添加到统计列表
    doctor_box_stats.append({
        '医生': doctor,
        '科室': unit,
        '新增卡号数量': new_cards_count,
        '总盒数': total_medicine_qty,
        '平均每卡盒数': round(total_medicine_qty / new_cards_count, 2) if new_cards_count > 0 else 0
    })

# 转换为DataFrame并按总盒数降序排序
doctor_box_stats_df = pd.DataFrame(doctor_box_stats).sort_values('总盒数', ascending=False)

# 创建一个ExcelWriter对象，用于输出Excel文件
with pd.ExcelWriter('6月新增盒装药品分析.xlsx', engine='openpyxl') as writer:
    # 1. 医生盒装药品统计（按总盒数降序）
    doctor_box_stats_df.to_excel(writer, sheet_name='医生盒装药品统计', index=False)
    
    # 2. 新增卡号盒装药品详细数据
    new_cards_box_data.to_excel(writer, sheet_name='新增盒装药品明细', index=False)
    
    # 3. 按医生分组的新增盒装药品明细
    for doctor, group in new_cards_box_data.groupby('医生'):
        sheet_name = f"{doctor}-盒装"[:31]  # Excel工作表名称最长31个字符
        group.to_excel(writer, sheet_name=sheet_name, index=False)

# 添加自动筛选功能
wb = load_workbook('6月新增盒装药品分析.xlsx')
for sheet_name in ['医生盒装药品统计', '新增盒装药品明细']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.auto_filter.ref = ws.dimensions

# 保存工作簿
wb.save('6月新增盒装药品分析.xlsx')

print('6月新增盒装药品分析报告已导出为 6月新增盒装药品分析.xlsx')
print(f"6月总新增卡号数: {len(new_cards_data)}")
print(f"6月新增盒装药品卡号数: {len(new_cards_box_data)}")
print(f"6月新增盒装药品总盒数: {new_cards_box_data['数量'].sum() if '数量' in new_cards_box_data.columns else 0}") 